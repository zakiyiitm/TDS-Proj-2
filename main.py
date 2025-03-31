from fastapi import FastAPI, Form, File, UploadFile  # type: ignore
from fastapi.responses import HTMLResponse  # type: ignore
from fastapi.middleware.cors import CORSMiddleware  # type: ignore
import os
import openpyxl  # type: ignore
from processing import fetch_answer
import re
import stat
import json
import base64
from io import BytesIO
from PIL import Image
import httpx  # type: ignore
import aiofiles
from typing import List
from git_api import GA1_13, GA2_3, GA2_7, GA4_8, GA2_9_file, GA2_6_file
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

app = FastAPI()

# CORS Configuration (Vercel allows any origin by default)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "tasks.xlsx")


def load_tasks_from_excel():
    if not os.path.exists(EXCEL_FILE):
        return {}
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    tasks = {row[0]: row[1] for row in sheet.iter_rows(
        min_row=2, values_only=True) if row[0] and row[1]}

    tasks_answers = {row[0]: row[2] for row in sheet.iter_rows(
        min_row=2, values_only=True) if row[0] and row[2]}
    workbook.close()
    return (tasks, tasks_answers)


TASKS, TASKS_ANSWERS = load_tasks_from_excel()


def classify_task(question: str) -> str:
    """Classify a question based on keyword matching with TASKS."""
    question_lower = question.lower()  # Convert to lowercase for case-insensitive matching
    for task_id, keyword in TASKS.items():
        if keyword.lower() in question_lower:
            return task_id  # Return the first matching task ID
    return "Unknown"  # Default if no match is found


def save_file(file: UploadFile):
    os.makedirs("uploads", exist_ok=True)
    if not file or not file.filename:
        return "Error: No file provided."
    # Define the file path
    file_path = os.path.join(os.getcwd(), "uploads", file.filename)
    try:
        # Write the file content manually
        with open(file_path, "wb") as buffer:
            buffer.write(file.file.read())
        # Set file permissions to 777 (read, write, execute for all)
        os.chmod(file_path, stat.S_IRWXU | stat.S_IRWXG | stat.S_IRWXO)
    except Exception as e:
        return f"Error saving file: {str(e)}"
    return file_path


def get_file_path(question: str) -> str:
    """Extracts a single filename from the question and returns its full path in the /uploads directory."""
    match = re.search(r'([^\/\\\s]+?\.[a-zA-Z0-9]+)', question)
    file = match.group(1) if match else None
    file_path = os.path.join(os.getcwd(), "uploads", file) if file else None
    return file_path if file_path and os.path.exists(file_path) else None


@app.get("/", response_class=HTMLResponse)
async def serve_form():
    file_path = os.path.join(os.path.dirname(__file__), "index.html")
    try:
        with open(file_path, "r") as file:
            return HTMLResponse(content=file.read())
    except FileNotFoundError:
        return HTMLResponse(content="<h1>index.html not found</h1>", status_code=404)


async def read_answer(task_id: str, question: str):
    print("reading from json")
    answer = TASKS_ANSWERS.get(task_id, "No answer found for this task.")
    return answer


def to_string(value):
    """Converts any type of value to a string representation."""
    if value is None:
        return "None"
    if not isinstance(value, str):
        try:
            # Converts lists, dicts, and serializable objects
            return json.dumps(value)
        except (TypeError, ValueError):
            return str(value)  # Fallback for other types
    return value


def Solve_Unknown_Task(question):
    BASE_URL = "https://aiproxy.sanand.workers.dev/openai/v1"
    data = {
        "model": "gpt-4o-mini",
        "messages": [{"role": "user", "content": question+" return only the answer"}]
    }
    API_KEY = os.getenv("AIPROXY_TOKEN")
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {API_KEY}"
    }
    response = httpx.post(BASE_URL + "/chat/completions",
                          json=data, headers=headers, timeout=60)

    return response.json().get("choices", [])[0].get("message", {}).get("content")


@app.post("/api/")
async def receive_question(question: str = Form(...), file: UploadFile = File(None)):
    # async def receive_question(question: str = Form(...), files: List[UploadFile] = File(None)):     # file = files[0]

    # if 'where is ' in question.lower():
    #     file_path = get_file_path(question)
    #     return {"question": question, "answer": file_path if file_path else "File not found"}

    task_id = classify_task(question)
    if task_id == "Unknown":
        print(question)
        answer = Solve_Unknown_Task(question)
    elif task_id in ['GA1.1']:
        answer = await read_answer(task_id=task_id, question=question)
    elif task_id in ['GA1.2', 'GA1.4', 'GA1.5', 'GA1.7', 'GA1.9', 'GA1.18']:
        answer = await fetch_answer(task_id=task_id, question=question, file_path="")
    elif task_id in ['GA1.3']:
        if file:
            print(file)
            if not os.getenv('VERCEL'):
                answer = await fetch_answer(task_id=task_id, question=question, file_path=file)
            else:
                answer = await read_answer(task_id=task_id, question=question)
        else:
            answer = await read_answer(task_id=task_id, question=question)
    elif task_id in ['GA1.16']:
        # print(os.getenv('VERCEL'))
        if file:
            print(file)
            answer = await fetch_answer(task_id=task_id, question=question, file_path=file)
        else:
            answer = await read_answer(task_id=task_id, question=question)
    elif task_id in ['GA1.8', 'GA1.10', 'GA1.12', 'GA1.14', 'GA1.15', 'GA1.17']:
        if file:
            print(file)
            answer = await fetch_answer(task_id=task_id, question=question, file_path=file)
        else:
            answer = await read_answer(task_id=task_id, question=question)
    elif task_id in ['GA1.6', 'GA1.11']:
        func_answer = ""
        if file:
            print(file)
            func_answer = await fetch_answer(task_id=task_id, question=question, file_path=file)
        answer = func_answer or await read_answer(task_id=task_id, question=question)
    elif task_id in ['GA1.13']:
        answer = GA1_13(question)
        # answer = "https://raw.githubusercontent.com/Telvinvarghese/Test/main/email.json"
    elif task_id in ['GA2.1']:
        answer = await read_answer(task_id=task_id, question=question)
    elif task_id in ['GA2.3']:
        answer = GA2_3(question)
        # answer = "https://telvinvarghese.github.io/website/"
    elif task_id in ['GA2.2', 'GA2.4']:
        if file:
            print(file)
            answer = await fetch_answer(task_id=task_id, question=question, file_path=file)
        else:
            answer = await read_answer(task_id=task_id, question=question)
    elif task_id in ['GA2.5']:
        if file:
            print(file)
            answer = await fetch_answer(task_id=task_id, question=question, file_path=file)
        else:
            answer = await fetch_answer(task_id=task_id, question=question, file_path="")
    elif task_id in ['GA2.6']:
        print(file)
        # file_content = await file.read()
        flag = await GA2_6_file(file)
        if flag == "True":
            answer = "https://api-git-main-telvinvargheses-projects.vercel.app/api"
        else:
            answer = "https://api-git-main-telvinvargheses-projects.vercel.app/api"
    elif task_id in ['GA2.7']:
        answer = GA2_7(question)
        # answer = "https://github.com/Telvinvarghese/Test"
    elif task_id in ['GA2.8']:
        answer = "https://hub.docker.com/repository/docker/telvinvarghese/py-hello/general"
    elif task_id in ['GA2.9']:
        print(file)
        # file_content = await file.read() 
        flag = await GA2_9_file(file)
        if flag == "True":
            answer = "https://tds-ga2-9.vercel.app/api"
        else:
            answer = "https://tds-ga2-9.vercel.app/api"
    elif task_id in ['GA2.10']:
        answer = "https://b45f-223-178-84-140.ngrok-free.app/"
    elif task_id in ["GA3.1", "GA3.2", "GA3.3", "GA3.5", "GA3.6"]:
        answer = await fetch_answer(task_id=task_id, question=question, file_path="")
    elif task_id in ["GA3.4"]:
        if file:
            print(file)
            answer = await fetch_answer(task_id=task_id, question=question, file_path=file)
        else:
            answer = await read_answer(task_id=task_id, question=question)
    elif task_id in ["GA3.7"]:
        answer = "https://tds-ga3-7.vercel.app/similarity"
    elif task_id in ["GA3.8"]:
        answer = "https://tds-ga3-8.vercel.app/execute"
    elif task_id in ['GA3.9']:
        answer = await read_answer(task_id=task_id, question=question)
    elif task_id in ['GA4.1', 'GA4.2', 'GA4.4', 'GA4.5', 'GA4.6', 'GA4.7']:
        answer = await fetch_answer(task_id=task_id, question=question, file_path="")
    elif task_id in ['GA4.3']:
        answer = "https://tds-ga4-3.vercel.app/api/outline"
    elif task_id in ['GA4.8']:
        answer = GA4_8(question)
        # answer = "https://github.com/Telvinvarghese/Test"
    elif task_id in ['GA4.9']:
        if file:
            print(file)
            answer = await fetch_answer(task_id=task_id, question=question, file_path=file)
        else:
            answer = await fetch_answer(task_id=task_id, question=question, file_path="")
    elif task_id in ['GA4.10']:
        answer = await read_answer(task_id=task_id, question=question)
    elif task_id in ['GA5.1', 'GA5.2', 'GA5.5', 'GA5.6', 'GA5.7']:
        if file:
            print(file)
            answer = await fetch_answer(task_id=task_id, question=question, file_path=file)
        else:
            answer = await read_answer(task_id=task_id, question=question)
    elif task_id in ['GA5.3', 'GA5.4']:
        if file:
            print(file)
            answer = await fetch_answer(task_id=task_id, question=question, file_path=file)
        else:
            answer = await fetch_answer(task_id=task_id, question=question, file_path="")
    elif task_id in ['GA5.8']:
        answer = await fetch_answer(task_id=task_id, question=question, file_path="")
    elif task_id in ['GA5.9']:
        answer = await fetch_answer(task_id=task_id, question=question, file_path="")
    elif task_id in ['GA5.10']:
        if file:
            print(file)
            answer = await fetch_answer(task_id=task_id, question=question, file_path=file)
            # image_url = f"data:image/png;base64,{answer}"
            # print(image_url)
            # img_data = base64.b64decode(answer)
            # img = Image.open(BytesIO(img_data))
            # img.show()
            # with open("reconstructed_image.png", "wb") as f:
            #     f.write(img_data)
    else:
        if file:
            # file_path = save_file(file)
            print(file)
            file_path = file
        answer = await read_answer(task_id=task_id, question=question)
        response = {"answer": answer}
        print(response)
        return response

    actual_answer = answer
    answer = to_string(answer)
    output = {"question": question, "task": task_id, "answer": answer,
              "file received": file.filename if file else "No file uploaded", }
    print("output :", output)
    print()
    response = {"answer": answer}
    print("response :", response)
    print()
    try:
        print("json output :", json.loads(answer))
    except json.JSONDecodeError:
        pass
    return response
